"""Bulk team invitations via CSV or Excel upload.

Two-stage flow:

  POST /api/team/invite/bulk/preview   (multipart file)
      → parses + validates CSV/XLSX against current tenant state, NO
        DB writes. Returns per-row status (ok | warning | error)
        plus a summary.

  POST /api/team/invite/bulk/commit    (JSON body of pre-validated rows)
      → re-checks cheap conditions, creates invitations one at a time,
        returns per-row result. Best-effort, no batch transaction.

Why two endpoints (rather than one with `dry_run=true`): cleaner OpenAPI,
easier to reason about, easier to test.

Format detection: filename extension first (.csv / .xlsx / .xls);
falls back to magic bytes (XLSX = PK ZIP header, anything else =
CSV). The parsed shape is identical so downstream validation
doesn't care which path the row came through.
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Iterable

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from pydantic import validate_email
from sqlalchemy.exc import IntegrityError

from app.models import Category, Invitation
from app.routes.deps import RequestContext, get_current_context
from app.schemas.invite_bulk import (
    BulkCommitInvitation,
    BulkCommitRequest,
    BulkCommitResponse,
    BulkCommitResultRow,
    BulkCommitSummary,
    BulkPreviewResponse,
    BulkPreviewRow,
    BulkPreviewSummary,
)
from app.services.invitations import (
    create_invitation as create_invitation_service,
)
from app.services.invitations import (
    has_live_invitation,
    is_already_member,
    send_invitation_email,
)

logger = logging.getLogger("app.team_bulk")

router = APIRouter()

CANONICAL_HEADERS = ["email", "name", "category"]
# Map every accepted header form (lowercased, accents stripped) to the
# canonical key we use internally. Lets admins use either English or
# Spanish CSVs — and tolerates common typos like the unaccented
# "categoria". Order is irrelevant because we read the file by header
# names, not column position.
HEADER_ALIASES = {
    "email": "email",
    "correo": "email",
    "name": "name",
    "nombre": "name",
    "category": "category",
    "categoria": "category",  # also matches "categoría" after accent strip
}
MAX_FILE_BYTES = 1 * 1024 * 1024  # 1 MB
MAX_ROWS = 5000

# CSV template returned by the GET endpoint and shipped to the frontend.
CSV_TEMPLATE = (
    "email,nombre,categoría\r\n"
    "maria@hospital.es,Maria Lopez,Adjunto\r\n"
    "juan@hospital.es,Juan Garcia,Residente R3\r\n"
)


def _strip_accents(s: str) -> str:
    """Lowercase, strip accents, used to normalize CSV header names so
    `categoría`, `categoria`, `Categoría` and `CATEGORÍA` all match the
    same canonical key."""
    import unicodedata

    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower())
        if unicodedata.category(c) != "Mn"
    )


def _require_admin(ctx: RequestContext) -> None:
    if "admin" not in ctx.membership.roles:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Admin role required"
        )


def _strip_bom(text: str) -> str:
    if text.startswith("﻿"):
        return text[1:]
    return text


def _detect_delimiter(text: str) -> str:
    """Sniff the CSV delimiter so semicolon-separated files (Excel ES/EU
    default) are accepted alongside comma-separated ones. Falls back to
    "," if the sample is too small or the sniffer fails.
    """
    # Look at the first ~2 KB; the first line is enough but Sniffer wants
    # a couple of rows for confidence.
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        # Sniffer choked (one-line file, weird quoting). Fall back to
        # whichever of `;` or `,` appears more in the first line.
        first_line = sample.splitlines()[0] if sample else ""
        return ";" if first_line.count(";") > first_line.count(",") else ","


def _validate_email(value: str) -> str | None:
    """Returns the normalized (lowercased) email or None if invalid."""
    try:
        # Pydantic's validate_email returns (name, email) tuple. Raises
        # PydanticCustomError on invalid input — we catch broadly because that
        # error class isn't part of the public API.
        _, normalized = validate_email(value)
        return normalized.lower()
    except Exception:
        return None


def _parse_csv(raw: bytes) -> list[dict[str, str]]:
    """Parse CSV bytes → list of rows as dicts. Raises HTTPException(400) on
    malformed header. Skips fully blank rows.
    """
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        raise HTTPException(
            status_code=400,
            detail="El archivo debe estar codificado en UTF-8.",
        )
    text = _strip_bom(text)

    # Spanish/European Excel defaults to ";" as the CSV separator because
    # the comma is used as a decimal separator. Detect by sniffing the
    # first non-empty line; fall back to "," if undetectable.
    delimiter = _detect_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        header = next(reader)
    except StopIteration:
        raise HTTPException(status_code=400, detail="El archivo CSV está vacío.")

    # Resolve each provided header to its canonical key via the alias map.
    # Unknown headers are reported. We require email + name; category is
    # optional in the data (admins can leave it blank) but the column
    # itself must be present so the mapping is unambiguous.
    raw_header = [(h or "").strip() for h in header]
    canonical: list[str | None] = []
    unknown: list[str] = []
    for h in raw_header:
        key = HEADER_ALIASES.get(_strip_accents(h))
        canonical.append(key)
        if key is None and h:
            unknown.append(h)
    if unknown:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Columnas no reconocidas: {', '.join(unknown)}. "
                "Cabeceras aceptadas: email, nombre, categoría "
                "(o sus equivalentes en inglés)."
            ),
        )
    missing = [k for k in CANONICAL_HEADERS if k not in canonical]
    if missing:
        # All three columns must be present. category may be empty per row,
        # but the column header itself is mandatory so each row dict has a
        # consistent shape.
        raise HTTPException(
            status_code=400,
            detail=(
                "Faltan columnas requeridas. La cabecera debe incluir: "
                "email, nombre, categoría."
            ),
        )

    rows: list[dict[str, str]] = []
    for raw_row in reader:
        # Skip blank lines (all empty / whitespace fields).
        if not raw_row or all((c or "").strip() == "" for c in raw_row):
            continue
        # Map each column by its canonical key (which respects the order
        # the admin's CSV used — columns can be in any order). Pad short
        # rows so trailing columns become empty strings instead of
        # IndexError.
        padded = (raw_row + [""] * len(canonical))[: len(canonical)]
        row: dict[str, str] = {k: "" for k in CANONICAL_HEADERS}
        for col_idx, key in enumerate(canonical):
            if key is not None:
                row[key] = (padded[col_idx] or "").strip()
        rows.append(row)
        if len(rows) > MAX_ROWS:
            raise HTTPException(
                status_code=400,
                detail=f"El archivo supera el máximo de {MAX_ROWS} filas.",
            )
    return rows


def _parse_xlsx(raw: bytes) -> list[dict[str, str]]:
    """Parse the first sheet of an .xlsx upload → list of row dicts
    keyed by canonical header names. Same return shape as
    `_parse_csv` so the calling code can dispatch on file type
    without caring about the parser branch downstream.

    We use openpyxl in read_only mode for predictable memory use
    on bigger files (5k-row cap still applies). Empty rows are
    skipped, just like the CSV path.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(
            filename=io.BytesIO(raw), read_only=True, data_only=True
        )
    except Exception as e:
        # openpyxl raises various exception types for malformed XLSX
        # (zipfile.BadZipFile, KeyError on missing parts, etc.) — flatten
        # them into a single user-facing 400.
        raise HTTPException(
            status_code=400,
            detail=f"No se pudo leer el archivo Excel: {e}",
        ) from None

    ws = wb.active
    if ws is None:
        raise HTTPException(
            status_code=400, detail="El archivo Excel no tiene hojas."
        )

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(
            status_code=400, detail="El archivo Excel está vacío."
        )

    # Excel cells come back as native Python types (int, float, datetime).
    # For the bulk-invite use case every column is text, so we coerce
    # to string and strip; None becomes "".
    raw_header = [(_cell_text(c)).strip() for c in header_row]
    canonical: list[str | None] = []
    unknown: list[str] = []
    for h in raw_header:
        key = HEADER_ALIASES.get(_strip_accents(h))
        canonical.append(key)
        if key is None and h:
            unknown.append(h)
    if unknown:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Columnas no reconocidas: {', '.join(unknown)}. "
                "Cabeceras aceptadas: email, nombre, categoría "
                "(o sus equivalentes en inglés)."
            ),
        )
    missing = [k for k in CANONICAL_HEADERS if k not in canonical]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Faltan columnas requeridas. La cabecera debe incluir: "
                "email, nombre, categoría."
            ),
        )

    rows: list[dict[str, str]] = []
    for raw_row in rows_iter:
        cells = [_cell_text(c) for c in raw_row]
        if not cells or all(c.strip() == "" for c in cells):
            continue
        padded = (cells + [""] * len(canonical))[: len(canonical)]
        row: dict[str, str] = {k: "" for k in CANONICAL_HEADERS}
        for col_idx, key in enumerate(canonical):
            if key is not None:
                row[key] = padded[col_idx].strip()
        rows.append(row)
        if len(rows) > MAX_ROWS:
            raise HTTPException(
                status_code=400,
                detail=f"El archivo supera el máximo de {MAX_ROWS} filas.",
            )
    return rows


def _cell_text(value) -> str:
    """Coerce an openpyxl cell value to a stripped string. Numbers
    survive as their `str()` form (rare in this dataset but harmless);
    None becomes "". We don't trim here — callers do."""
    if value is None:
        return ""
    return str(value)


def _is_xlsx_upload(file: UploadFile, raw: bytes) -> bool:
    """Tell whether this upload should be routed to the Excel parser.
    Filename extension is the primary signal — admins overwhelmingly
    leave Excel-default extensions in place. Magic bytes (`PK\\x03\\x04`,
    the ZIP header that XLSX files inherit) are the fallback for
    browsers that strip the filename, paranoid clients, etc."""
    name = (file.filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return True
    if name.endswith(".csv"):
        return False
    # No extension → look at the first 4 bytes.
    return raw[:4] == b"PK\x03\x04"


def _load_categories_by_lower_name(ctx: RequestContext) -> dict[str, Category]:
    cats = ctx.db.query(Category).filter(Category.tenant_id == ctx.tenant.id).all()
    return {c.name.strip().lower(): c for c in cats}


@router.get("/team/invite/bulk/template")
def download_template():
    from fastapi.responses import Response

    return Response(
        content=CSV_TEMPLATE,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="plantilla-equipo.csv"'
        },
    )


@router.post(
    "/team/invite/bulk/preview",
    response_model=BulkPreviewResponse,
)
async def preview_bulk(
    file: UploadFile = File(...),
    ctx: RequestContext = Depends(get_current_context),
) -> BulkPreviewResponse:
    _require_admin(ctx)

    raw = await file.read()
    if len(raw) > MAX_FILE_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"El archivo supera el máximo de {MAX_FILE_BYTES // 1024} KB.",
        )

    rows = _parse_xlsx(raw) if _is_xlsx_upload(file, raw) else _parse_csv(raw)
    cats_by_name = _load_categories_by_lower_name(ctx)

    out: list[BulkPreviewRow] = []
    seen_emails: set[str] = set()
    valid = warning = errored = 0

    for idx, r in enumerate(rows, start=1):
        email_raw = r["email"]
        name_raw = r["name"]
        cat_raw = r["category"]

        # 1) email format
        normalized_email = _validate_email(email_raw) if email_raw else None
        # 2) name
        name_clean = name_raw.strip()

        error: str | None = None
        warning_msg: str | None = None
        category_id: int | None = None

        if not email_raw:
            error = "El email es obligatorio."
        elif not normalized_email:
            error = "Email no válido."
        elif not name_clean:
            error = "El nombre es obligatorio."
        elif len(name_clean) > 255:
            error = "El nombre supera los 255 caracteres."
        else:
            # Category lookup
            if cat_raw:
                cat = cats_by_name.get(cat_raw.lower())
                if not cat:
                    error = f"La categoría «{cat_raw}» no existe en este equipo."
                else:
                    category_id = cat.id

        if error is None and normalized_email is not None:
            # 3) duplicate within file
            if normalized_email in seen_emails:
                error = "Email duplicado en el archivo."
            else:
                seen_emails.add(normalized_email)

        if error is None and normalized_email is not None:
            # 4) already a member?
            if is_already_member(ctx.db, ctx.tenant.id, normalized_email):
                error = "Esta persona ya es miembro del equipo."

        if error is None and normalized_email is not None:
            # 5) live invitation already exists → warning, not error
            if has_live_invitation(ctx.db, ctx.tenant.id, normalized_email):
                warning_msg = (
                    "Ya existe una invitación pendiente; se reemplazará al confirmar."
                )

        status_code = (
            "error" if error else ("warning" if warning_msg else "ok")
        )
        if status_code == "ok":
            valid += 1
        elif status_code == "warning":
            warning += 1
        else:
            errored += 1

        out.append(
            BulkPreviewRow(
                row_number=idx,
                email=normalized_email or email_raw,
                name=name_clean,
                category=cat_raw or None,
                category_id=category_id,
                status=status_code,
                error=error,
                warning=warning_msg,
            )
        )

    return BulkPreviewResponse(
        rows=out,
        summary=BulkPreviewSummary(
            total_rows=len(out),
            valid_rows=valid,
            warning_rows=warning,
            error_rows=errored,
        ),
    )


@router.post(
    "/team/invite/bulk/commit",
    response_model=BulkCommitResponse,
)
def commit_bulk(
    payload: BulkCommitRequest,
    ctx: RequestContext = Depends(get_current_context),
) -> BulkCommitResponse:
    _require_admin(ctx)

    if len(payload.rows) > MAX_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"El número de filas supera el máximo de {MAX_ROWS}.",
        )

    results: list[BulkCommitResultRow] = []
    committed = skipped = errored = 0
    seen: set[str] = set()

    for row in payload.rows:
        email_norm = _validate_email(row.email)
        if not email_norm:
            errored += 1
            results.append(
                BulkCommitResultRow(
                    row_number=row.row_number,
                    email=row.email,
                    status="error",
                    reason="Email no válido.",
                )
            )
            continue

        if email_norm in seen:
            skipped += 1
            results.append(
                BulkCommitResultRow(
                    row_number=row.row_number,
                    email=email_norm,
                    status="skipped",
                    reason="Email duplicado en la solicitud.",
                )
            )
            continue
        seen.add(email_norm)

        # Validate category belongs to this tenant if supplied (defensive).
        if row.category_id is not None:
            cat = ctx.db.get(Category, row.category_id)
            if not cat or cat.tenant_id != ctx.tenant.id:
                errored += 1
                results.append(
                    BulkCommitResultRow(
                        row_number=row.row_number,
                        email=email_norm,
                        status="error",
                        reason="Categoría desconocida.",
                    )
                )
                continue

        if is_already_member(ctx.db, ctx.tenant.id, email_norm):
            skipped += 1
            results.append(
                BulkCommitResultRow(
                    row_number=row.row_number,
                    email=email_norm,
                    status="skipped",
                    reason="Ya es miembro del equipo.",
                )
            )
            continue

        try:
            created = create_invitation_service(
                ctx.db,
                tenant_id=ctx.tenant.id,
                email=email_norm,
                person_name=row.name.strip(),
                created_by_membership_id=ctx.membership.id,
                category_id=row.category_id,
                roles=["member"],
            )
        except IntegrityError:
            ctx.db.rollback()
            errored += 1
            results.append(
                BulkCommitResultRow(
                    row_number=row.row_number,
                    email=email_norm,
                    status="error",
                    reason="Conflicto al crear la invitación.",
                )
            )
            continue
        except Exception as exc:  # pragma: no cover - defensive
            logger.exception(
                "Unexpected error creating bulk invitation tenant=%s email=%s: %s",
                ctx.tenant.slug,
                email_norm,
                exc,
            )
            ctx.db.rollback()
            errored += 1
            results.append(
                BulkCommitResultRow(
                    row_number=row.row_number,
                    email=email_norm,
                    status="error",
                    reason="Error inesperado.",
                )
            )
            continue

        # Best-effort email; failure must NOT abort the batch.
        # Skipped entirely when the caller passes send_email=False
        # (onboarding bulk-import: defer the email blast until the
        # admin triggers per-row "Enviar invitación" later).
        if payload.send_email:
            try:
                send_invitation_email(ctx.db, tenant_id=ctx.tenant.id, created=created)
            except Exception as exc:  # pragma: no cover - send_invitation_email already swallows
                logger.error(
                    "Bulk invite email failure tenant=%s email=%s err=%s",
                    ctx.tenant.slug,
                    email_norm,
                    exc,
                )

        committed += 1
        results.append(
            BulkCommitResultRow(
                row_number=row.row_number,
                email=email_norm,
                status="ok",
                invitation=BulkCommitInvitation(
                    id=created.invitation.id,
                    email=created.invitation.email,
                    expires_at=created.invitation.expires_at,
                    accept_url=created.accept_url,
                ),
            )
        )

    logger.warning(
        "Bulk invite commit tenant=%s committed=%s skipped=%s errored=%s",
        ctx.tenant.slug,
        committed,
        skipped,
        errored,
    )

    return BulkCommitResponse(
        results=results,
        summary=BulkCommitSummary(
            committed=committed,
            skipped=skipped,
            errored=errored,
        ),
    )


# Silence unused import warnings if the linter is fussy
_ = (Iterable, Invitation)
